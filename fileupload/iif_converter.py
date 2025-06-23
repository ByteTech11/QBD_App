import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import re
from openpyxl.utils import get_column_letter

def convert_excel_to_iif(excel_file, output_iif_path):
    def get_merged_cell_value(sheet, row, col):
        cell_coord = f"{get_column_letter(col)}{row}"
        for merged_range in sheet.merged_cells.ranges:
            if cell_coord in merged_range:
                min_col, min_row, max_col, max_row = merged_range.bounds
                return sheet.cell(row=min_row, column=min_col).value
        return sheet.cell(row=row, column=col).value

    wb = load_workbook(excel_file, data_only=True)
    sheet = wb['Sheet1']

    emp_type_header_cell = get_merged_cell_value(sheet, 5, 12) 
    print(f"DEBUG: Emp Type header from L5: {repr(emp_type_header_cell)}")

    if emp_type_header_cell is not None:
        full_header = str(emp_type_header_cell).strip().upper()
        match = re.search(r'\b(SK|ON|IS)\b', full_header)
        if match:
            emp_type_header = match.group(1)
        else:
            emp_type_header = None
    else:
        emp_type_header = None

    print(f"DEBUG: Extracted Emp Type header from L5: '{emp_type_header}'")

    try:
        df = pd.read_excel(excel_file, engine='openpyxl', header=13, usecols="F:S")
    except Exception as e:
        raise RuntimeError(f"Error reading Excel file: {e}")

    df.columns = df.columns.str.strip().str.replace('\n', ' ').str.replace('\r', ' ')
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

    if 'Stat Hours' not in df.columns:
        df['Stat Hours'] = 0

    date_found = None
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=5, column=col).value
        if isinstance(val, str):
            matches = re.findall(r'\d{2}-[A-Za-z]{3}-\d{2,4}', val)
            if len(matches) >= 2:
                date_found = matches[1]
                break
            elif len(matches) == 1 and date_found is None:
                date_found = matches[0]

    if not date_found:
        raise RuntimeError("No valid date found in row 5")

    extracted_date = datetime.strptime(date_found, '%d-%b-%y')
    adjusted_date = extracted_date - timedelta(days=2)
    formatted_date = adjusted_date.strftime('%m/%d/%Y')

    df_cleaned = df.dropna(subset=['Employee', 'Emp Num'])

    iif_header = "!TIMEACT\tDATE\tJOB\tEMP\tITEM\tPITEM\tDURATION\tPROJ\tNOTE\tXFERTOPAYROLL\tBILLINGSTATUS\n"
    iif_data = [iif_header]

    max_hours_map = {'IS': 48, 'SK': 80, 'ON': 88}
    default_max_hours = 88

    for idx, row in df_cleaned.iterrows():
        emp_name = row['Employee']
        emp_num = str(int(row['Emp Num'])) if not pd.isna(row['Emp Num']) else ""
        emp_type_raw = row['Emp Type'] if pd.notna(row['Emp Type']) else ""
        if emp_type_raw.strip():
            emp_type = emp_type_raw.strip().upper()
            print(f"DEBUG: Row {idx} uses Emp Type from row: '{emp_type}'")
        else:
            emp_type = emp_type_header if emp_type_header else 'ON'
            print(f"DEBUG: Row {idx} Emp Type missing, using header value: '{emp_type}'")
        emp = f"{emp_name}{{{emp_num}}}"


        reg_hours = row['Reg Hours'] if 'Reg Hours' in row and pd.notna(row['Reg Hours']) else 0
        stat_hours = row['Stat Hours'] if pd.notna(row['Stat Hours']) else 0
        ph_hours = row['PH Hours'] if 'PH Hours' in row and pd.notna(row['PH Hours']) else 0

        max_hours = max_hours_map.get(emp_type, default_max_hours)
        if reg_hours > max_hours:
            reg_hours = max_hours
        note = f"E_NUM:{emp_num},EMP_TYPE:{emp_type}"
        if reg_hours > 0:
            iif_data.append(
                f"TIMEACT\t{formatted_date}\tDefault Job\t{emp}\tST Rate\tRegular Pay\t{reg_hours}\t\t{note}\tY\t1"
            )
        if ph_hours > 0:
            iif_data.append(
                f"TIMEACT\t{formatted_date}\tDefault Job\t{emp}\tPH Hours\tPublic Holiday\t{ph_hours}\t\t{note}\tY\t1"
            )
        if stat_hours > 0:
            iif_data.append(
                f"TIMEACT\t{formatted_date}\tDefault Job\t{emp}\tSTAT Hours\tStat Holiday\t{stat_hours}\t\t{note}\tY\t1"
            )
    iif_content = "\n".join(iif_data)

    with open(output_iif_path, 'w') as f:
        f.write(iif_content)
    print(f"IIF file successfully created: {output_iif_path}")
    return output_iif_path